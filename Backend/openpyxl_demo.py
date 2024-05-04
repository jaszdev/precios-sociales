from openpyxl import Workbook

wb = Workbook()

# Active worksheet
ws = wb.active

# Create Header
ws['A1'] = 'Producto'
ws['B1'] = 'Código'
ws['C1'] = 'Clasificación'
ws['D1'] = 'Precio Económico'
ws['E1'] = 'Factor A'
ws['F1'] = 'Factor B'
ws['G1'] = 'Precio Social'

# First Row
ws['A2'] = "Cemento"
ws['B2'] = 'NP090 - Cemento, cal y yeso'
ws['C2'] = 'No transable'
ws['D2'] = '1000'
ws['E2'] = '3.13'
ws['F2'] = '2.56'
ws['G2'] = '=(D2*E2)/F2'

wb.save('demo.xlsx')